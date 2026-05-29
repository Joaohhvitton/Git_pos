"""
=============================================================
  AUTOMAÇÃO DE VERSÕES POS — TMS7 + SGV
=============================================================
"""

import asyncio
import shutil
from pathlib import Path
from datetime import datetime

import zipfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from playwright.async_api import async_playwright


# ──────────────────────────────────────────────────────────
#  CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────
_DIR         = Path(__file__).parent

def _achar_base(pasta):
    for nome in ["modelo_parking", "parking_version"]:
        for ext in [".xlsx", ".xlsm", ".xls"]:
            p = pasta / f"{nome}{ext}"
            if p.exists():
                return p
    return pasta / "modelo_parking.xlsx"

BASE_ARQUIVO = _achar_base(_DIR)
DOWNLOAD_DIR = _DIR / "downloads"
BACKUP_DIR   = _DIR / "backups"

TMS7 = {
    "url"    : "https://app.ger7.com.br/TMS7/RedeFlexProd/Pages/Login.aspx"
               "?ReturnUrl=%2fTMS7%2fRedeFlexProd",
    "usuario": "joao.delima",
    "senha"  : "rede@2026!",
    "sheet"  : "TMS7-ParqueFinanceira",
}

SGV = {
    "url"    : "https://sgv.redetendencia.com.br/sgv/login/autenticacao.jsp",
    "usuario": "fabio.tarso",
    "senha"  : "rede@12345",
    "sheet"  : "SGV-ParqueRecarga",
}

HEADLESS = False
MODO_DEBUG = True


# ──────────────────────────────────────────────────────────
#  REGRAS VISUAIS — extraídas do modelo_parking.xlsx
# ──────────────────────────────────────────────────────────

# Nome das tabelas Excel em cada aba (TableStyleMedium2)
TMS7_TABLE_NAME = "Tabela26"
SGV_TABLE_NAME  = "Tabela3"
TABLE_STYLE     = "TableStyleMedium2"

# Larguras de coluna (em unidades Excel) por nome de cabeçalho — TMS7
COLWIDTHS_TMS7 = {
    "Número":                   11.44,
    "Cód. Estabelecimento":     22.11,
    "Estabelecimento":          40.33,
    "CNPJ":                     14.44,
    "Cidade":                   28.44,
    "Parametrização Atual":     22.11,
    "Param. Programada":         8.00,
    "Id Primário":              24.44,
    "Modelo":                   11.11,
    "Versão da Aplicação":      20.66,
    "Versão telecarga":         19.89,
    "Última comunicação":       21.00,
    "Tipo de comunicação":      21.55,
    "Serial Simcard":           20.44,
    "IMSI":                      8.89,
    "Operadora":                13.44,
    "Status":                   10.11,
    "Data da Criação":          17.55,
    "Data Ativação":            16.66,
    "Última comunicação RDS7":  25.55,
    "Rede":                     11.33,
    "Working key de dados":     22.11,
    "Chave PIX":                13.00,
    "MCC":                       9.11,
    "Nome do comprovante":      40.33,
    "Sub":                       8.44,
    "Teste":                     9.55,
    "Adquirido":                12.89,
    "Intregração":               8.66,
    "Condicional":               8.00,
    "VersãoRecarga":             8.00,
    "Versão TEM":                8.00,
}

# Larguras de coluna (em unidades Excel) por nome de cabeçalho — SGV
COLWIDTHS_SGV = {
    "Adquirente":                   12.44,
    "Alocado?":                     10.55,
    "Numero série":                 14.55,
    "Status terminal":              16.00,
    "Tipo terminal":                14.11,
    "Fabricante":                   11.89,
    "Modelo":                       25.44,
    "Data negociacao":              16.89,
    "Data alocação":                14.66,
    "Versão S.O. Terminal":         20.55,
    "Versão Binário":               15.11,
    "Protocolo terminal":           18.89,
    "Tipo proprietário":            17.11,
    "Proprietário":                 13.11,
    "Data última venda":            18.33,
    "Iccid última venda":           18.44,
    "Código Empresa":               16.89,
    "Operadora última venda":       23.44,
    "Empresa":                      10.44,
    "Protocolo última venda":       22.55,
    "Código ERP":                   12.89,
    "Data Iccid inventário":        20.44,
    "Situação":                     10.33,
    "Iccid inventário":             16.11,
    "Tipo pessoa":                  13.11,
    "Operadora Iccid inventário":   25.55,
    "Tipo cliente":                 13.00,
    "Canais venda":                 14.44,
    "Código estabelecimento":       24.00,
    "Estabelecimento":              17.44,
    "Cep":                           8.00,
    "Empresa Estabelecimento Id":   27.44,
    "Logradouro":                   12.55,
    "Empresa Estabelecimento":      25.44,
    "Numero":                        9.89,
    "Mando":                         8.00,
    "Complemento":                  15.33,
    "Bairro":                        8.00,
    "Cidade":                        9.00,
    "UF":                            8.00,
    "Bloqueado":                    12.00,
    "Data cadastro":                14.55,
    "Consultor":                    11.44,
    "Supervisor":                   12.11,
    "Última telecarga":             16.89,
    "APN":                           8.00,
    "Terminal":                     10.44,
    "Coluna1":                       8.00,
}


# ──────────────────────────────────────────────────────────
#  UTILITÁRIOS
# ──────────────────────────────────────────────────────────
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def preparar_pastas():
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)


def fazer_backup():
    if BASE_ARQUIVO.exists():
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = BACKUP_DIR / f"Base_backup_{ts}.xlsx"
        shutil.copy2(BASE_ARQUIVO, dest)
        log(f"✅ Backup criado → {dest.name}")


async def screenshot(page, nome):
    caminho = DOWNLOAD_DIR / f"screenshot_{nome}_{datetime.now().strftime('%H%M%S')}.png"
    await page.screenshot(path=str(caminho))
    log(f"📸 Screenshot → {caminho.name}")


# ──────────────────────────────────────────────────────────
#  AUTOMAÇÃO TMS7
# ──────────────────────────────────────────────────────────
async def baixar_tms7(context):
    log("─" * 50)
    log("📥 TMS7 — iniciando...")
    page = await context.new_page()

    try:
        await page.goto(TMS7["url"], wait_until="networkidle")
        await page.wait_for_timeout(1000)

        await page.fill("#ctl00_cphPrincipal_txtLogin", TMS7["usuario"])
        await page.fill("#ctl00_cphPrincipal_txtSenha", TMS7["senha"])
        await page.keyboard.press("Enter")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(1500)
        log("✅ Login TMS7")

        await page.click("#ctl00_wclMenuCustom_Terminais > a")
        await page.wait_for_timeout(800)

        await page.click("#ctl00_wclMenuCustom_ul_Terminais > li:nth-child(3) > a")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(1500)
        log("✅ Terminais > Busca")

        clicou = await page.evaluate("""() => {
            const elementos = document.querySelectorAll('*');
            for (const el of elementos) {
                if (el.children.length === 0 && el.textContent.trim() === 'Status') {
                    const container = el.closest('p') || el.closest('td') || el.parentElement;
                    if (container) {
                        const btn = container.querySelector('a[role="button"], a.ui-button, a.custom-combobox-toggle');
                        if (btn) { btn.click(); return true; }
                    }
                }
            }
            return false;
        }""")

        if clicou:
            log("✅ Seta do Status clicada")
        else:
            log("⚠️  Seta do Status não encontrada via JS — tentando por índice")
            await page.locator("a.custom-combobox-toggle").nth(4).click()

        await page.wait_for_timeout(600)

        clicou_todos = await page.evaluate("""() => {
            const itens = document.querySelectorAll('ul.ui-autocomplete li, ul[id^="ui-id"] li');
            for (const li of itens) {
                if (li.textContent.trim() === 'Todos' && li.offsetParent !== null) {
                    li.click(); return true;
                }
            }
            return false;
        }""")

        if not clicou_todos:
            await page.locator("ul.ui-autocomplete li:visible, .ui-menu-item:visible").first.click()

        await page.wait_for_timeout(500)
        log("✅ Status = Todos")

        await page.click("#ctl00_cphPrincipal_btnFiltro")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(2000)
        log("✅ Filtro aplicado")

        log("⏳ Exportando XLS...")
        async with page.expect_download(timeout=120_000) as dl_info:
            await page.click("#ctl00_cphPrincipal_btnExportar")
            try:
                await page.wait_for_selector("div.ui-dialog button", timeout=4000)
                await page.click("div.ui-dialog button:first-child")
                log("✅ Popup confirmado")
            except Exception:
                pass

        download = await dl_info.value
        sufixo = Path(download.suggested_filename).suffix or ".xls"
        dest   = DOWNLOAD_DIR / f"tms7_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufixo}"
        await download.save_as(str(dest))
        log(f"✅ TMS7 salvo → {dest.name}")
        return dest

    except Exception as e:
        log(f"❌ Erro TMS7: {e}")
        await screenshot(page, "tms7_erro")
        return None
    finally:
        await page.close()


# ──────────────────────────────────────────────────────────
#  AUTOMAÇÃO SGV
# ──────────────────────────────────────────────────────────
async def baixar_sgv(context):
    log("─" * 50)
    log("📥 SGV — iniciando...")
    page = await context.new_page()

    try:
        await page.goto(SGV["url"], wait_until="networkidle")
        await page.wait_for_timeout(1000)

        await page.fill("#usrNome", SGV["usuario"])
        await page.fill("#password", SGV["senha"])
        await page.keyboard.press("Enter")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(2000)
        log("✅ Login SGV")

        async def js_click_menu(texto, espera_navegacao=False):
            js = """(texto) => {
                const links = document.querySelectorAll('li a');
                for (const a of links) {
                    if (a.textContent.trim() === texto) {
                        a.click();
                        return true;
                    }
                }
                return false;
            }"""
            if espera_navegacao:
                try:
                    async with page.expect_navigation(wait_until="networkidle", timeout=15000):
                        await page.evaluate(js, texto)
                except Exception:
                    await page.wait_for_load_state("networkidle")
            else:
                try:
                    await page.evaluate(js, texto)
                except Exception:
                    pass
                await page.wait_for_timeout(1000)

        await js_click_menu("Relatórios", espera_navegacao=False)
        await page.wait_for_timeout(800)
        log("✅ Clicou em Relatórios")

        await js_click_menu("Administrativo", espera_navegacao=False)
        await page.wait_for_timeout(800)
        log("✅ Clicou em Administrativo")

        await js_click_menu("Terminais", espera_navegacao=True)
        await page.wait_for_timeout(2000)
        log("✅ Clicou em Terminais")

        await screenshot(page, "sgv_filtros")

        abriu = False
        for sel in [
            "text=Selecione as empresas",
            "span:has-text('Selecione as empresas')",
            "div:has-text('Selecione as empresas')",
            "[title='Selecione as empresas']",
        ]:
            try:
                await page.locator(sel).first.click(timeout=5000)
                log("✅ Dropdown de empresas aberto")
                abriu = True
                break
            except Exception:
                continue

        await page.wait_for_timeout(800)
        await page.wait_for_timeout(500)

        search_box = page.locator("input[type='text']").last
        try:
            bbox = await search_box.bounding_box(timeout=5000)
            if bbox:
                x = bbox["x"] - 15
                y = bbox["y"] + bbox["height"] / 2
                await page.mouse.click(x, y)
                log(f"✅ Clicou no quadrado vazio (x={x:.0f}, y={y:.0f})")
        except Exception as e:
            log(f"⚠️  Clique por coordenada falhou: {e}")

        await page.wait_for_timeout(500)
        await screenshot(page, "sgv_empresas_selecionadas")

        await page.keyboard.press("Escape")
        await page.wait_for_timeout(500)

        for sel in ["select[name*='status']", "select[id*='status']",
                    "select[name*='Status']", "select[id*='Status']"]:
            try:
                loc = page.locator(sel).first
                await loc.wait_for(state="visible", timeout=3000)
                for valor in [{"label": "Todos"}, {"label": "TODOS"},
                               {"value": "T"}, {"value": ""}]:
                    try:
                        await loc.select_option(**valor)
                        log("✅ Status = Todos")
                        break
                    except Exception:
                        continue
                break
            except Exception:
                continue

        await page.wait_for_timeout(500)

        log("⏳ Clicando em Gerar...")
        for sel in ["input[value='Gerar']", "input[value='GERAR']",
                    "button:has-text('Gerar')", "a:has-text('Gerar')",
                    "input[type='submit']"]:
            try:
                await page.click(sel, timeout=4000)
                break
            except Exception:
                continue

        log("⏳ Aguardando dialog de confirmação...")
        try:
            await page.locator("button.botaoSim").last.wait_for(state="visible", timeout=15_000)
            await page.wait_for_timeout(600)
            await page.locator("button.botaoSim").last.click(force=True)
            log("✅ Clicou Sim")
            try:
                await page.wait_for_load_state("networkidle", timeout=15_000)
            except Exception:
                await page.wait_for_timeout(3000)
        except Exception as e:
            log(f"⚠️  Problema no Sim: {e}")

        await page.wait_for_timeout(1500)

        async def relogar_sgv_se_necessario():
            if "login" in page.url.lower() or "autenticacao" in page.url.lower():
                log("🔄 Sessão expirou — fazendo re-login no SGV...")
                await page.fill("#usrNome", SGV["usuario"])
                await page.fill("#password", SGV["senha"])
                await page.keyboard.press("Enter")
                await page.wait_for_load_state("networkidle")
                await page.wait_for_timeout(1500)
                await page.goto(
                    page.url.replace("autenticacao.jsp", "").rstrip("/")
                    .rsplit("/sgv", 1)[0]
                    + "/sgv/paginas/relatorio/acompanhamento/acompanhamentoRelatorio.jsf",
                    wait_until="networkidle"
                )
                await page.wait_for_timeout(1500)
                log("✅ Re-login feito e acompanhamento reaberto")

        await relogar_sgv_se_necessario()
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(2000)
        log("✅ Tela de acompanhamento de relatórios aberta")

        dest = None
        for tentativa in range(20):
            log(f"🔄 Pesquisando status... (tentativa {tentativa+1}/20)")
            try:
                await page.click("input[value='Pesquisar'], button:has-text('Pesquisar')", timeout=5000)
                await page.wait_for_load_state("networkidle")
                await page.wait_for_timeout(1000)
            except Exception:
                pass

            primeira_linha_status = await page.evaluate("""
                () => {
                    const linhas = document.querySelectorAll('table tbody tr, tr.resultado, tr[class*="row"]');
                    for (const tr of linhas) {
                        const tds = tr.querySelectorAll('td');
                        if (tds.length < 3) continue;
                        const texto = tr.textContent;
                        if (!texto.includes('TERMINAIS')) continue;
                        for (let i = tds.length - 1; i >= 0; i--) {
                            const t = tds[i].textContent.trim();
                            if (t) return t;
                        }
                    }
                    return 'não encontrado';
                }
            """)
            log(f"   Status: {primeira_linha_status}")

            processado = any(s in primeira_linha_status.lower()
                             for s in ["processado", "download", "concluído", "concluido"])

            if processado:
                log("✅ Relatório pronto! Clicando no ícone CSV...")
                await page.wait_for_timeout(500)
                try:
                    async with page.expect_download(timeout=60_000) as dl_info:
                        clicou = await page.evaluate("""
                            () => {
                                const linhas = document.querySelectorAll('table tbody tr, tr');
                                for (const tr of linhas) {
                                    if (!tr.textContent.includes('TERMINAIS')) continue;
                                    const links = tr.querySelectorAll('a');
                                    for (const a of links) {
                                        const img = a.querySelector('img');
                                        const href = a.href || '';
                                        const src  = img ? img.src : '';
                                        if (src.toLowerCase().includes('csv') ||
                                            href.toLowerCase().includes('csv') ||
                                            (img && img.title && img.title.toLowerCase().includes('csv'))) {
                                            a.click();
                                            return true;
                                        }
                                    }
                                    if (links[0]) { links[0].click(); return 'first-link'; }
                                }
                                return false;
                            }
                        """)
                        if not clicou:
                            log("⚠️  Ícone CSV não encontrado via JS — tentando seletores")
                            await page.click("img[src*='csv'], a[href*='csv'], td.acoes a:first-child", timeout=5000)

                    download = await dl_info.value
                    sufixo = Path(download.suggested_filename).suffix or ".zip"
                    dest   = DOWNLOAD_DIR / f"sgv_{datetime.now().strftime('%Y%m%d_%H%M%S')}{sufixo}"
                    await download.save_as(str(dest))
                    log(f"✅ SGV baixado → {dest.name}")
                    break
                except Exception as e:
                    log(f"❌ Erro ao baixar: {e}")
                    await screenshot(page, "sgv_download_erro")
                    break
            else:
                log(f"   Aguardando 15s...")
                await page.wait_for_timeout(15_000)

        if not dest:
            await screenshot(page, "sgv_timeout")
            log("❌ Relatório não ficou pronto a tempo — veja screenshot sgv_timeout")

        return dest

    except Exception as e:
        log(f"❌ Erro SGV: {e}")
        await screenshot(page, "sgv_erro")
        return None
    finally:
        await page.close()


# ──────────────────────────────────────────────────────────
#  LEITURA DE RELATÓRIOS
def _limpar_df(df):
    """
    Converte colunas com numeros grandes (CNPJ, Id Primario, Serial, ICCID)
    para string, evitando notacao cientifica ao gravar no Excel.
    """
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(
                lambda v: v[1:] if isinstance(v, str) and v.startswith("'") else v
            )
            continue
        try:
            sample = df[col].dropna()
            if len(sample) == 0:
                continue
            max_abs = sample.abs().max()
            if max_abs >= 1e9:
                def _to_str(v, _col=col):
                    if pd.isna(v):
                        return None
                    iv = int(v)
                    return str(iv) if float(iv) == float(v) else str(v)
                df[col] = df[col].apply(_to_str)
        except Exception:
            pass
    return df


# ──────────────────────────────────────────────────────────
def ler_relatorio(caminho):
    try:
        sufixo = caminho.suffix.lower()

        if sufixo == ".zip":
            with zipfile.ZipFile(caminho, "r") as z:
                csvs = [f for f in z.namelist() if f.lower().endswith(".csv")]
                if not csvs:
                    log(f"❌ Nenhum CSV encontrado dentro do ZIP: {caminho.name}")
                    return None
                csv_nome = csvs[0]
                log(f"   📂 Extraindo: {csv_nome}")
                with z.open(csv_nome) as f:
                    for sep in [";", ",", "\t"]:
                        try:
                            import io
                            dados = f.read()
                            df = pd.read_csv(
                                io.BytesIO(dados),
                                sep=sep,
                                encoding="latin-1",
                                on_bad_lines="skip",
                                skiprows=5,
                                dtype=str,
                            )
                            if len(df.columns) > 1:
                                log(f"   ✅ CSV lido: {len(df)} linhas, {len(df.columns)} colunas (sep='{sep}')")
                                return _limpar_df(df)
                        except Exception:
                            continue
            log(f"❌ Não foi possível ler o CSV dentro do ZIP")
            return None

        if sufixo == ".xls":
            df = pd.read_excel(caminho, engine="xlrd", dtype=str)
        else:
            df = pd.read_excel(caminho, engine="openpyxl", dtype=str)
        return _limpar_df(df)

    except Exception as e:
        log(f"❌ Erro ao ler {caminho.name}: {e}")
        return None


# ──────────────────────────────────────────────────────────
#  FORMATAÇÃO VISUAL — aplica as regras do modelo na planilha
# ──────────────────────────────────────────────────────────
COLUNAS_FORMULA_TMS7 = {"Intregração", "Condicional", "VersãoRecarga"}
COLUNAS_VALOR_TMS7   = {"Teste", "Adquirido"}
COLUNAS_EXTRA_TMS7   = COLUNAS_FORMULA_TMS7 | COLUNAS_VALOR_TMS7 | {"Versão TEM", "Integração"}


def _aplicar_formatacao_visual(ws, colwidths, table_name, n_rows):
    """
    Aplica as regras visuais do modelo_parking.xlsx na aba:
      1. Redimensiona a Tabela Excel para cobrir exatamente os dados gravados.
      2. Garante TableStyleMedium2 (mesmo estilo do modelo).
      3. Aplica as larguras de coluna originais por nome de cabeçalho.
      4. Garante que a linha 1 (cabeçalho) está visível e com altura padrão.
    """
    from openpyxl.utils import get_column_letter

    # 1. Redimensiona a Tabela Excel
    if table_name in ws.tables:
        tbl = ws.tables[table_name]
        max_col = ws.max_column
        nova_ref = f"A1:{get_column_letter(max_col)}{n_rows + 1}"
        tbl.ref = nova_ref
        if tbl.tableStyleInfo:
            tbl.tableStyleInfo.name = TABLE_STYLE
        log(f"   📐 Tabela '{table_name}' → {nova_ref} ({TABLE_STYLE})")
    else:
        log(f"   ⚠️  Tabela '{table_name}' não encontrada na aba")

    # 2. Larguras de coluna por cabeçalho
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    aplicadas = 0
    for col_idx, header in enumerate(headers, start=1):
        if header and header in colwidths:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = colwidths[header]
            aplicadas += 1
    log(f"   📏 Larguras aplicadas: {aplicadas}/{len(headers)} colunas")

    # 3. Altura da linha de cabeçalho (padrão modelo = 15)
    ws.row_dimensions[1].height = 15


# ──────────────────────────────────────────────────────────
#  ATUALIZAÇÃO DA BASE  (xlwings — preserva Tabelas Dinâmicas)
# ──────────────────────────────────────────────────────────

def _colunas_e_ultima_linha(ws):
    headers_raw = ws.range("1:1").value
    if isinstance(headers_raw, str):
        headers_raw = [headers_raw]
    headers = [h for h in (headers_raw or []) if h is not None]
    ultima = ws.range("A1").current_region.last_cell.row
    return headers, ultima


def _salvar_valores_manuais(ws, headers, ultima_linha):
    if ultima_linha < 2:
        return {}, {}
    if "Número" not in headers:
        return {}, {}

    idx_num  = headers.index("Número")   + 1
    idx_tst  = headers.index("Teste")    + 1 if "Teste"     in headers else None
    idx_adq  = headers.index("Adquirido") + 1 if "Adquirido" in headers else None

    def _ler_col(idx):
        if idx is None:
            return []
        vals = ws.range((2, idx), (ultima_linha, idx)).value
        if vals is None:
            return []
        return vals if isinstance(vals, list) else [vals]

    numeros = _ler_col(idx_num)
    testes  = _ler_col(idx_tst)
    adqs    = _ler_col(idx_adq)

    def _flat(lst):
        return [v[0] if isinstance(v, list) else v for v in lst]

    numeros = _flat(numeros)
    testes  = _flat(testes)
    adqs    = _flat(adqs)

    saved_teste = dict(zip(numeros, testes))
    saved_adq   = dict(zip(numeros, adqs))

    qtd_t = sum(1 for v in saved_teste.values() if v not in (None, ""))
    qtd_a = sum(1 for v in saved_adq.values()   if v not in (None, ""))
    log(f"   💾 Teste preenchido: {qtd_t:,}  |  Adquirido: {qtd_a:,}")
    return saved_teste, saved_adq


def _escrever_tms7(wb_xlwings, wb_openpyxl, df_novos):
    ws = wb_xlwings.sheets[TMS7["sheet"]]
    headers, ultima = _colunas_e_ultima_linha(ws)
    cols_df = list(df_novos.columns)

    saved_teste, saved_adq = _salvar_valores_manuais(ws, headers, ultima)

    idx_num_df = cols_df.index("Número") if "Número" in cols_df else None

    log(f"   ✏️  Preparando {len(df_novos):,} linhas...")
    dados = []
    for row in df_novos.itertuples(index=False, name=None):
        numero = row[idx_num_df] if idx_num_df is not None else None
        linha = []
        for h in headers:
            if h in COLUNAS_FORMULA_TMS7 or h == "Versão TEM":
                linha.append(None)
            elif h == "Teste":
                linha.append(saved_teste.get(numero))
            elif h == "Adquirido":
                linha.append(saved_adq.get(numero))
            elif h == "Integração":
                linha.append(None)
            elif h in cols_df:
                linha.append(row[cols_df.index(h)])
            else:
                linha.append(None)
        dados.append(linha)

    if ultima > 1:
        log(f"   🗑️  Removendo {ultima - 1:,} linhas antigas...")
        last_col_letter = ws.range((1, len(headers))).get_address(False, False).split(":")[0].rstrip("0123456789")
        ws.range(f"A2:{last_col_letter}{ultima}").clear()

    log(f"   💾 Gravando {len(dados):,} linhas na planilha...")
    ws.range("A2").value = dados

    formulas_tms7 = {
        "Intregração": (
            '=IF(ISERROR(VLOOKUP([@CNPJ],POSIntegrada!B:B,1,0)),"-","MegaGiro")'
        ),
        "Condicional": (
            '=IF([@Sub]&[@Intregração]="RedeFlexMegaGiro","MegaGiro",'
            'IF([@Sub]&[@Intregração]="SubTal-","SubTal","RedeFlex"))'
        ),
        "VersãoRecarga": (
            "=IFERROR(IFERROR(_xlfn.XLOOKUP([@[Id Primário]],"
            "'SGV-ParqueRecarga'!C:C,'SGV-ParqueRecarga'!K:K),"
            "_xlfn.XLOOKUP([@[Id Primário]],"
            "'TEM-Parque'!B:B,'TEM-Parque'!E:E)),\"\")"
        ),
    }
    n = len(dados)
    for col_nome, formula in formulas_tms7.items():
        if col_nome in headers:
            col_idx = headers.index(col_nome) + 1
            if n > 0:
                ws.range((2, col_idx), (n + 1, col_idx)).formula = formula
                log(f"   📐 Fórmula '{col_nome}' → {n:,} linhas")

    # ── Aplica formatação visual do modelo ──────────────────
    log("   🎨 Aplicando formatação visual (modelo)...")
    ws_op = wb_openpyxl[TMS7["sheet"]]
    _aplicar_formatacao_visual(ws_op, COLWIDTHS_TMS7, TMS7_TABLE_NAME, n)

    log(f"✅ TMS7 → {n:,} registros gravados")


def _escrever_sgv(wb_xlwings, wb_openpyxl, df_novos):
    ws = wb_xlwings.sheets[SGV["sheet"]]
    headers, ultima = _colunas_e_ultima_linha(ws)
    cols_df = list(df_novos.columns)
    n = len(df_novos)

    dados = []
    for row in df_novos.itertuples(index=False, name=None):
        linha = []
        for h in headers:
            val = row[cols_df.index(h)] if h in cols_df else None
            # Remove apostrofe-prefixo do Excel nos campos de série/ICCID
            if isinstance(val, str) and val.startswith("'"):
                val = val[1:]
            linha.append(val)
        dados.append(linha)

    if ultima > 1:
        last_col_letter = ws.range((1, len(headers))).get_address(False, False).split(":")[0].rstrip("0123456789")
        ws.range(f"A2:{last_col_letter}{ultima}").clear()

    log(f"   💾 Gravando {n:,} linhas SGV...")
    ws.range("A2").value = dados

    # ── Aplica formatação visual do modelo ──────────────────
    log("   🎨 Aplicando formatação visual (modelo)...")
    ws_op = wb_openpyxl[SGV["sheet"]]
    _aplicar_formatacao_visual(ws_op, COLWIDTHS_SGV, SGV_TABLE_NAME, n)

    log(f"✅ SGV → {n:,} registros gravados")


# ──────────────────────────────────────────────────────────
#  ORQUESTRAÇÃO DA ATUALIZAÇÃO
# ──────────────────────────────────────────────────────────

def _refresh_sincronamente_e_salvar(wb_xlwings, app):
    """
    Atualiza conexoes/pivots e salva de forma SINCRONA, evitando
    o OLE error 0x800ac472.

    Causa raiz: RefreshAll() com BackgroundQuery=True dispara threads
    separadas no Excel, mantendo o COM objeto perpetuamente ocupado.
    Mesmo CalculateUntilAsyncQueriesDone() nao resolve porque o Excel
    usa um mecanismo de lock diferente para essas threads.

    Solucao definitiva:
      1. Desativa BackgroundQuery em todas as conexoes e PivotCaches
         -> o refresh passa a ser SINCRONAMENTE bloqueante
      2. Chama RefreshAll() — Python espera terminar antes de continuar
      3. Salva via wb.api.Save() (COM direto, sem DisplayAlerts)
      4. Restaura BackgroundQuery=True nas conexoes
    """
    import time as _time

    api = wb_xlwings.api

    # ── Etapa 1: desativa background refresh em todas as conexoes ──
    log("\u2699\ufe0f  Desativando background refresh nas conexoes...")
    conexoes_alteradas = []
    try:
        for conn in api.Connections:
            try:
                conn.OLEDBConnection.BackgroundQuery = False
                conexoes_alteradas.append(("OLEDB", conn))
            except Exception:
                pass
            try:
                conn.ODBCConnection.BackgroundQuery = False
                conexoes_alteradas.append(("ODBC", conn))
            except Exception:
                pass
        log(f"   {len(conexoes_alteradas)} conexao(oes) configurada(s) para refresh sincrono")
    except Exception as e:
        log(f"   \u26a0\ufe0f  Conexoes: {e}")

    pivot_caches_alterados = []
    try:
        for pc in api.PivotCaches():
            try:
                pc.BackgroundQuery = False
                pivot_caches_alterados.append(pc)
            except Exception:
                pass
        log(f"   {len(pivot_caches_alterados)} PivotCache(s) configurado(s) para refresh sincrono")
    except Exception as e:
        log(f"   \u26a0\ufe0f  PivotCaches: {e}")

    # ── Etapa 2: refresh sincrono (bloqueia ate terminar) ──
    log("\U0001f504 Executando RefreshAll() sincrono...")
    try:
        api.RefreshAll()
        log("   \u2705 RefreshAll() concluido")
    except Exception as e:
        log(f"   \u26a0\ufe0f  RefreshAll: {e}")

    _time.sleep(2)  # margem de seguranca

    # ── Etapa 2b: atualiza cada PivotTable explicitamente ──
    log("🔄 Atualizando tabelas dinâmicas (PivotTables)...")
    total_pivots = 0
    for sheet in wb_xlwings.sheets:
        try:
            pts = sheet.api.PivotTables()
            for i in range(1, pts.Count + 1):
                try:
                    pt = pts.Item(i)
                    pt.PivotCache().Refresh()
                    pt.RefreshTable()
                    total_pivots += 1
                    log(f"   ✅ Pivot '{pt.Name}' em '{sheet.name}' atualizado")
                except Exception as ep:
                    log(f"   ⚠️  Pivot {i} em '{sheet.name}': {ep}")
        except Exception:
            pass
    if total_pivots == 0:
        log("   ℹ️  Nenhuma PivotTable encontrada")
    else:
        log(f"   ✅ {total_pivots} PivotTable(s) atualizadas")
    _time.sleep(1)  # margem extra apos pivots

    # ── Etapa 3: salva via COM direto ──
    log("\U0001f4be Salvando arquivo...")
    api.Save()
    log("   \u2705 Arquivo salvo com sucesso")

    # ── Etapa 4: restaura background refresh ──
    for tipo, conn in conexoes_alteradas:
        try:
            if tipo == "OLEDB":
                conn.OLEDBConnection.BackgroundQuery = True
            else:
                conn.ODBCConnection.BackgroundQuery = True
        except Exception:
            pass
    for pc in pivot_caches_alterados:
        try:
            pc.BackgroundQuery = True
        except Exception:
            pass
    log("\u2705 Background refresh restaurado")


def atualizar_base(arq_tms7, arq_sgv):
    global BASE_ARQUIVO
    log("─" * 50)
    log("📊 Atualizando base Excel (xlwings + openpyxl)...")
    fazer_backup()

    if not BASE_ARQUIVO.exists():
        log(f"❌ Base não encontrada: {BASE_ARQUIVO}")
        log("   ➡️  Coloque a base na mesma pasta que o automacao.py")
        return

    try:
        import xlwings as xw
    except ImportError:
        log("❌ xlwings não instalado. Execute:  pip install xlwings")
        return

    import time
    import tempfile

    # ── Copia para caminho temporário sem acentos ──────────────────────────────
    # O COM do Excel invisível falha em pastas com acentos (ex: "Automação").
    # Solução: trabalhar em %TEMP% com nome simples e copiar de volta no final.
    tmp_xlsx = Path(tempfile.gettempdir()) / "automacao_base_tmp.xlsx"
    try:
        shutil.copy2(BASE_ARQUIVO, tmp_xlsx)
        log(f"   📋 Cópia temporária → {tmp_xlsx.name}")
    except PermissionError:
        log("❌ Arquivo aberto no Excel — feche-o e tente novamente.")
        return
    except Exception as e:
        log(f"❌ Erro ao copiar base: {e}")
        return

    app = xw.App(visible=True, add_book=False)
    try:
        # Desativa Protected View e segurança de automação (evita falha no Open invisível)
        app.api.AutomationSecurity = 1  # msoAutomationSecurityLow
        app.api.DisplayAlerts = False
        wb_xlwings = app.books.open(str(tmp_xlsx), update_links=False)
        app.api.Visible = False  # esconde após abrir com sucesso

        df_tms7 = None
        df_sgv  = None

        if arq_tms7 and arq_tms7.exists():
            df_tms7 = ler_relatorio(arq_tms7)

        if arq_sgv and arq_sgv.exists():
            df_sgv = ler_relatorio(arq_sgv)

        # ── Fase 1: xlwings escreve dados e fórmulas ──
        if df_tms7 is not None:
            _escrever_tms7_dados(wb_xlwings, df_tms7)

        if df_sgv is not None:
            _escrever_sgv_dados(wb_xlwings, df_sgv)

        log("🔄 Atualizando tabelas dinâmicas e fórmulas...")
        _refresh_sincronamente_e_salvar(wb_xlwings, app)
        log("✅ Dados gravados na cópia temporária.")
        wb_xlwings.close()

        # ── Fase 2: openpyxl aplica formatação visual ──
        log("🎨 Aplicando formatação visual com openpyxl...")
        wb_op = load_workbook(str(tmp_xlsx))

        if df_tms7 is not None:
            ws_op = wb_op[TMS7["sheet"]]
            _aplicar_formatacao_visual(ws_op, COLWIDTHS_TMS7, TMS7_TABLE_NAME, len(df_tms7))

        if df_sgv is not None:
            ws_op = wb_op[SGV["sheet"]]
            _aplicar_formatacao_visual(ws_op, COLWIDTHS_SGV, SGV_TABLE_NAME, len(df_sgv))

        wb_op.save(str(tmp_xlsx))
        wb_op.close()

        # ── Copia o resultado de volta para o destino original ──
        shutil.copy2(tmp_xlsx, BASE_ARQUIVO)
        log(f"✅ Arquivo atualizado → {BASE_ARQUIVO.name}")

    except Exception as e:
        log(f"❌ Erro ao atualizar base: {e}")
        import traceback
        log(traceback.format_exc())
    finally:
        try:
            app.quit()
        except Exception:
            pass
        try:
            tmp_xlsx.unlink(missing_ok=True)
        except Exception:
            pass



# ── Funções auxiliares separadas (dados via xlwings) ──────

def _escrever_tms7_dados(wb_xlwings, df_novos):
    """Escreve dados e fórmulas via xlwings (sem formatação visual)."""
    ws = wb_xlwings.sheets[TMS7["sheet"]]
    headers, ultima = _colunas_e_ultima_linha(ws)
    cols_df = list(df_novos.columns)

    saved_teste, saved_adq = _salvar_valores_manuais(ws, headers, ultima)
    idx_num_df = cols_df.index("Número") if "Número" in cols_df else None

    log(f"   ✏️  Preparando {len(df_novos):,} linhas (TMS7)...")
    dados = []
    for row in df_novos.itertuples(index=False, name=None):
        numero = row[idx_num_df] if idx_num_df is not None else None
        linha = []
        for h in headers:
            if h in COLUNAS_FORMULA_TMS7 or h == "Versão TEM":
                linha.append(None)
            elif h == "Teste":
                linha.append(saved_teste.get(numero))
            elif h == "Adquirido":
                linha.append(saved_adq.get(numero))
            elif h == "Integração":
                linha.append(None)
            elif h in cols_df:
                linha.append(row[cols_df.index(h)])
            else:
                linha.append(None)
        dados.append(linha)

    if ultima > 1:
        log(f"   🗑️  Removendo {ultima - 1:,} linhas antigas...")
        last_col_letter = ws.range((1, len(headers))).get_address(False, False).split(":")[0].rstrip("0123456789")
        ws.range(f"A2:{last_col_letter}{ultima}").clear()

    log(f"   💾 Gravando {len(dados):,} linhas na planilha...")
    ws.range("A2").value = dados

    n = len(dados)

    # ── Resize da tabela usando ws.api.Range() (mais confiavel que .api do xlwings) ──
    if n > 0:
        try:
            import re as _re
            tabela = ws.api.ListObjects(TMS7_TABLE_NAME)
            # Extrai a ultima coluna do range atual da tabela (ex: "AF" de "$A$1:$AF$14929")
            tbl_addr = tabela.Range.Address.replace('$', '')
            m = _re.match(r'([A-Z]+)\d+:([A-Z]+)\d+', tbl_addr)
            last_tbl_col = m.group(2) if m else "AF"
            nova_area = ws.api.Range(f"A1:{last_tbl_col}{n + 1}")
            tabela.Resize(nova_area)
            log(f"   📏 Tabela '{TMS7_TABLE_NAME}' redimensionada → A1:{last_tbl_col}{n + 1}")
        except Exception as e:
            log(f"   ⚠️  Resize da tabela: {e}")

    # ── Fórmulas com referências diretas de célula (independem do resize da tabela) ──
    # Usa D2, Z2, AC2 etc. — o Excel ajusta automaticamente para cada linha da coluna
    if n > 0:
        from openpyxl.utils import get_column_letter as _gcl

        def _col(nome):
            return _gcl(headers.index(nome) + 1) if nome in headers else None

        c_cnpj   = _col("CNPJ")
        c_sub    = _col("Sub")
        c_intreg = _col("Intregração")
        c_idprim = _col("Id Primário")

        formulas_tms7 = {}

        if c_cnpj:
            formulas_tms7["Intregração"] = (
                f'=IF(ISERROR(VLOOKUP({c_cnpj}2,POSIntegrada!B:B,1,0)),"-","MegaGiro")'
            )
        if c_sub and c_intreg:
            formulas_tms7["Condicional"] = (
                f'=IF({c_sub}2&{c_intreg}2="RedeFlexMegaGiro","MegaGiro",'
                f'IF({c_sub}2&{c_intreg}2="SubTal-","SubTal","RedeFlex"))'
            )
        if c_idprim:
            formulas_tms7["VersãoRecarga"] = (
                f"=IFERROR(IFERROR(_xlfn.XLOOKUP({c_idprim}2,"
                f"'SGV-ParqueRecarga'!C:C,'SGV-ParqueRecarga'!K:K),"
                f"_xlfn.XLOOKUP({c_idprim}2,"
                f"'TEM-Parque'!B:B,'TEM-Parque'!E:E)),\"\")"
            )

        for col_nome, formula in formulas_tms7.items():
            if col_nome in headers:
                col_idx = headers.index(col_nome) + 1
                ws.range((2, col_idx), (n + 1, col_idx)).formula = formula
                log(f"   📐 Fórmula '{col_nome}' → {n:,} linhas")

    log(f"✅ TMS7 dados → {n:,} registros")

def _escrever_sgv_dados(wb_xlwings, df_novos):
    """Escreve dados via xlwings (sem formatação visual)."""
    ws = wb_xlwings.sheets[SGV["sheet"]]
    headers, ultima = _colunas_e_ultima_linha(ws)
    cols_df = list(df_novos.columns)
    n = len(df_novos)

    dados = []
    for row in df_novos.itertuples(index=False, name=None):
        linha = []
        for h in headers:
            val = row[cols_df.index(h)] if h in cols_df else None
            # Remove apostrofe-prefixo do Excel nos campos de série/ICCID
            if isinstance(val, str) and val.startswith("'"):
                val = val[1:]
            linha.append(val)
        dados.append(linha)

    if ultima > 1:
        last_col_letter = ws.range((1, len(headers))).get_address(False, False).split(":")[0].rstrip("0123456789")
        ws.range(f"A2:{last_col_letter}{ultima}").clear()

    log(f"   💾 Gravando {n:,} linhas SGV...")
    ws.range("A2").value = dados
    log(f"✅ SGV dados → {n:,} registros")


# ──────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────
def ultimo_arquivo(pasta: Path, prefixo: str):
    arquivos = sorted(pasta.glob(f"{prefixo}_*"), reverse=True)
    return arquivos[0] if arquivos else None


async def main():
    print()
    print("=" * 55)
    print("  AUTOMAÇÃO DE VERSÕES POS — TMS7 + SGV")
    print(f"  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}")
    print("=" * 55)

    preparar_pastas()

    if MODO_DEBUG:
        arq_tms7 = ultimo_arquivo(DOWNLOAD_DIR, "tms7")
        arq_sgv  = ultimo_arquivo(DOWNLOAD_DIR, "sgv")
        if arq_tms7:
            log(f"🐛 DEBUG — reutilizando TMS7: {arq_tms7.name}")
        else:
            log("🐛 DEBUG — nenhum arquivo TMS7 encontrado em downloads/")
        if arq_sgv:
            log(f"🐛 DEBUG — reutilizando SGV: {arq_sgv.name}")
        else:
            log("🐛 DEBUG — nenhum arquivo SGV encontrado em downloads/")
    else:
        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=HEADLESS,
                args=["--start-maximized"],
            )
            context = await browser.new_context(
                accept_downloads=True,
                viewport={"width": 1366, "height": 768},
            )

            arq_tms7 = await baixar_tms7(context)
            arq_sgv  = await baixar_sgv(context)

            await browser.close()

    atualizar_base(arq_tms7, arq_sgv)

    print()
    print("=" * 55)
    print("  ✅  CONCLUÍDO!")
    print("=" * 55)


if __name__ == "__main__":
    asyncio.run(main())